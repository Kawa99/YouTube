import io
import json
import zipfile
from datetime import datetime

import pytest
from openpyxl import load_workbook

import routes
from app import create_app
from models import (
    AffiliateProductEvidence,
    Asset,
    Channel,
    ChannelDerivedSummary,
    ChannelHistory,
    ChannelLabel,
    ChannelSnapshot,
    CollectionRun,
    ContentThesis,
    Experiment,
    ExperimentCheckpoint,
    OwnedAnalyticsCredential,
    OwnedVideoAnalytics,
    PackagingExperiment,
    RedTeamReview,
    RetentionDiagnostic,
    SponsorEvidence,
    ThesisEvidence,
    ThesisMonetizationMap,
    ThesisScore,
    ThesisTopic,
    Video,
    VideoAsset,
    VideoDerivedMetric,
    VideoDisclosure,
    VideoHistory,
    VideoLabel,
    VideoLabelAudit,
    VideoMetadataChange,
    VideoRightsChecklist,
    VideoSnapshot,
    db,
)


@pytest.fixture
def client(monkeypatch):
    monkeypatch.setenv("DATABASE_URL", "sqlite:///:memory:")

    app = create_app()
    app.config["TESTING"] = True

    with app.app_context():
        db.drop_all()
        db.create_all()
        with app.test_client() as test_client:
            yield test_client
        db.session.remove()
        db.drop_all()


def test_api_data_empty(client):
    response = client.get("/api/data")

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["counts"]["total_videos"] == 0


def test_api_data_pagination_limits(client):
    with client.application.app_context():
        channels = [
            Channel(channel_username=f"@channel_{idx}", subscribers=idx)
            for idx in range(30)
        ]
        db.session.add_all(channels)
        db.session.commit()

    response = client.get("/api/data?page=1&limit=10")

    assert response.status_code == 200
    payload = response.get_json()
    assert len(payload["channels"]["items"]) == 10

    pagination = payload["channels"]["pagination"]
    assert pagination["total_items"] == 30
    assert pagination["total_pages"] == 3
    assert pagination["current_page"] == 1
    assert pagination["has_next"] is True


def test_api_data_invalid_parameters_fallback(client):
    response = client.get("/api/data?page=-5&limit=invalid_string")

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["query"]["page"] == 1
    assert payload["query"]["limit"] == 25


def test_api_data_includes_and_sorts_by_engagement_rate(client):
    with client.application.app_context():
        channel = Channel(channel_username="@engagement_sort_channel", subscribers=2000)
        db.session.add(channel)
        db.session.flush()

        high_engagement = Video(
            youtube_video_id="engagement_sort_high",
            title="High engagement",
            views=100,
            likes=8,
            comments=2,
            channel_id=channel.id,
        )
        low_engagement = Video(
            youtube_video_id="engagement_sort_low",
            title="Low engagement",
            views=100,
            likes=1,
            comments=1,
            channel_id=channel.id,
        )
        zero_views = Video(
            youtube_video_id="engagement_sort_zero",
            title="Zero views",
            views=0,
            likes=999,
            comments=999,
            channel_id=channel.id,
        )
        db.session.add_all([high_engagement, low_engagement, zero_views])
        db.session.commit()

    response = client.get("/api/data?sort_column=engagement_rate&sort_direction=desc")

    assert response.status_code == 200
    payload = response.get_json()
    items = payload["videos"]["items"]
    assert payload["query"]["sort_column"] == "engagement_rate"
    assert len(items) >= 3
    assert "engagement_rate" in items[0]
    assert items[0]["youtube_video_id"] == "engagement_sort_high"
    assert items[0]["engagement_rate"] == 10.0
    assert items[-1]["youtube_video_id"] == "engagement_sort_zero"
    assert items[-1]["engagement_rate"] == 0.0


def test_api_data_sorts_by_like_and_comment_rate(client):
    with client.application.app_context():
        channel = Channel(channel_username="@rate_sort_channel", subscribers=1500)
        db.session.add(channel)
        db.session.flush()

        highest_like_rate = Video(
            youtube_video_id="like_rate_high",
            title="Highest like rate",
            views=100,
            likes=20,
            comments=1,
            channel_id=channel.id,
        )
        highest_comment_rate = Video(
            youtube_video_id="comment_rate_high",
            title="Highest comment rate",
            views=100,
            likes=1,
            comments=30,
            channel_id=channel.id,
        )
        baseline = Video(
            youtube_video_id="rate_baseline",
            title="Baseline",
            views=100,
            likes=2,
            comments=2,
            channel_id=channel.id,
        )
        db.session.add_all([highest_like_rate, highest_comment_rate, baseline])
        db.session.commit()

    like_rate_response = client.get(
        "/api/data?sort_column=like_rate&sort_direction=desc"
    )
    assert like_rate_response.status_code == 200
    like_rate_items = like_rate_response.get_json()["videos"]["items"]
    assert like_rate_items[0]["youtube_video_id"] == "like_rate_high"
    assert like_rate_items[0]["like_rate"] == 20.0

    comment_rate_response = client.get(
        "/api/data?sort_column=comment_rate&sort_direction=desc"
    )
    assert comment_rate_response.status_code == 200
    comment_rate_items = comment_rate_response.get_json()["videos"]["items"]
    assert comment_rate_items[0]["youtube_video_id"] == "comment_rate_high"
    assert comment_rate_items[0]["comment_rate"] == 30.0


def test_api_data_paginates_large_video_fixture(client):
    with client.application.app_context():
        channel = Channel(channel_username="@large_fixture_channel", subscribers=1000)
        db.session.add(channel)
        db.session.flush()
        db.session.bulk_save_objects(
            [
                Video(
                    youtube_video_id=f"large_fixture_{index}",
                    title=f"Large fixture {index}",
                    views=index,
                    likes=index % 100,
                    comments=index % 10,
                    channel_id=channel.id,
                )
                for index in range(10_000)
            ]
        )
        db.session.commit()

    response = client.get(
        "/api/data?page=2&limit=25&sort_column=views&sort_direction=desc"
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["videos"]["pagination"]["total_items"] == 10_000
    assert payload["videos"]["pagination"]["current_page"] == 2
    assert len(payload["videos"]["items"]) == 25
    assert payload["videos"]["items"][0]["views"] == 9974
    assert payload["videos"]["items"][-1]["views"] == 9950


def test_single_video_scraper_displays_engagement_rates(client, monkeypatch):
    monkeypatch.setattr(routes, "YOUTUBE_API_KEY", "test-api-key")
    monkeypatch.setattr(
        routes,
        "get_video_data",
        lambda _video_id: {
            "youtube_video_id": "dQw4w9WgXcQ",
            "channel_username": "@test_channel",
            "subscribers": "1000",
            "title": "Single video test",
            "views": "10000",
            "likes": "500",
            "comments": "200",
            "posted": "2025-01-01",
            "video_length": "0:05:00",
            "transcript": "Test transcript",
            "description": "Test description",
        },
    )

    response = client.post(
        "/",
        data={"video_url": "https://www.youtube.com/watch?v=dQw4w9WgXcQ"},
    )

    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "Engagement Rate" in body
    assert "Like Rate" in body
    assert "Comment Rate" in body
    assert "5.00%" in body
    assert "2.00%" in body
    assert "7.00%" in body


def test_export_csv_success(client):
    response = client.get("/export?format=csv")

    assert response.status_code == 200
    assert response.mimetype == "text/csv"
    body = response.get_data(as_text=True)
    assert "=== VIDEOS ===" in body
    assert "=== CHANNELS ===" in body
    assert "=== VIDEO_SNAPSHOTS ===" in body
    assert "=== CHANNEL_SNAPSHOTS ===" in body
    assert "=== VIDEO_METADATA_CHANGES ===" in body
    assert "=== VIDEO_LABELS ===" in body
    assert "=== VIDEO_LABEL_AUDITS ===" in body
    assert "=== CHANNEL_LABELS ===" in body
    assert "=== COLLECTION_RUNS ===" in body
    assert "=== VIDEO_DERIVED_METRICS ===" in body
    assert "=== CHANNEL_DERIVED_SUMMARIES ===" in body
    assert "=== PACKAGING_EXPERIMENTS ===" in body
    assert "=== CONTENT_THESES ===" in body
    assert "=== THESIS_EVIDENCE ===" in body
    assert "=== THESIS_TOPICS ===" in body
    assert "=== THESIS_SCORES ===" in body
    assert "=== RED_TEAM_REVIEWS ===" in body
    assert "=== THESIS_MONETIZATION_MAPS ===" in body
    assert "=== SPONSOR_EVIDENCE ===" in body
    assert "=== AFFILIATE_PRODUCT_EVIDENCE ===" in body
    assert "=== ASSETS ===" in body
    assert "=== VIDEO_ASSETS ===" in body
    assert "=== VIDEO_RIGHTS_CHECKLISTS ===" in body
    assert "=== VIDEO_DISCLOSURES ===" in body
    assert "=== OWNED_ANALYTICS_CREDENTIALS ===" in body
    assert "=== OWNED_VIDEO_ANALYTICS ===" in body
    assert "=== RETENTION_DIAGNOSTICS ===" in body
    assert "=== EXPERIMENTS ===" in body
    assert "=== EXPERIMENT_CHECKPOINTS ===" in body


def test_export_xlsx_success(client):
    response = client.get("/export?format=xlsx")

    assert response.status_code == 200
    assert (
        response.mimetype
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(io.BytesIO(response.data), read_only=True)
    try:
        assert set(workbook.sheetnames) == {
            "videos",
            "channels",
            "channel_snapshots",
            "video_snapshots",
            "video_metadata_changes",
            "video_labels",
            "video_label_audits",
            "channel_labels",
            "collection_runs",
            "video_derived_metrics",
            "channel_derived_summaries",
            "packaging_experiments",
            "content_theses",
            "thesis_evidence",
            "thesis_topics",
            "thesis_scores",
            "red_team_reviews",
            "thesis_monetization_maps",
            "sponsor_evidence",
            "affiliate_product_evidence",
            "assets",
            "video_assets",
            "video_rights_checklists",
            "video_disclosures",
            "owned_analytics_credentials",
            "owned_video_analytics",
            "retention_diagnostics",
            "experiments",
            "experiment_checkpoints",
            "channel_videos",
            "channel_history",
            "video_history",
            "video_metadata_history",
        }
    finally:
        workbook.close()


def test_research_zip_export_contains_schema_files_and_filters(client):
    with client.application.app_context():
        channel = Channel(
            channel_username="@finance_channel",
            subscribers=1000,
            youtube_channel_id="UC_FINANCE",
            channel_name="Finance Channel",
            subscriber_count=1000,
        )
        other_channel = Channel(
            channel_username="@gaming_channel",
            subscribers=500,
            youtube_channel_id="UC_GAMING",
            channel_name="Gaming Channel",
            subscriber_count=500,
        )
        db.session.add_all([channel, other_channel])
        db.session.flush()

        video = Video(
            youtube_video_id="finance_video",
            youtube_channel_id="UC_FINANCE",
            title="Finance Outlier",
            views=10000,
            likes=500,
            comments=100,
            duration_seconds=600,
            channel_id=channel.id,
            published_at=datetime(2026, 1, 1, 0, 0, 0),
        )
        other_video = Video(
            youtube_video_id="gaming_video",
            youtube_channel_id="UC_GAMING",
            title="Gaming Baseline",
            views=100,
            likes=5,
            comments=1,
            duration_seconds=600,
            channel_id=other_channel.id,
            published_at=datetime(2026, 1, 1, 0, 0, 0),
        )
        db.session.add_all([video, other_video])
        db.session.flush()

        run = CollectionRun(
            run_type="channel_uploads",
            status="completed",
            input_type="channel_id",
            input_value="UC_FINANCE",
            requested_limit=50,
            quota_estimate=4,
            items_found=1,
            items_saved=1,
            items_failed=0,
        )
        db.session.add(run)
        db.session.flush()

        db.session.add_all(
            [
                VideoLabel(
                    video_id=video.id,
                    niche="finance",
                    format="explainer",
                    faceless_status="faceless",
                    review_status="reviewed",
                ),
                VideoLabel(
                    video_id=other_video.id,
                    niche="gaming",
                    format="commentary",
                    faceless_status="faceless",
                    review_status="reviewed",
                ),
                ChannelLabel(
                    channel_id=channel.id,
                    primary_niche="finance",
                    primary_format="explainer",
                    faceless_status="faceless",
                ),
                ChannelSnapshot(
                    channel_id=channel.id,
                    subscriber_count=1000,
                    view_count=50000,
                    video_count=12,
                    collection_run_id=run.id,
                ),
                VideoSnapshot(
                    video_id=video.id,
                    view_count=10000,
                    like_count=500,
                    comment_count=100,
                    subscriber_count_at_snapshot=1000,
                    collection_run_id=run.id,
                ),
                VideoDerivedMetric(
                    video_id=video.id,
                    snapshot_at=datetime(2026, 1, 2, 0, 0, 0),
                    age_days=1,
                    views_per_day=10000,
                    views_per_subscriber=10,
                    channel_recent_median_views=1000,
                    relative_performance=10,
                    duration_bucket="8-15m",
                    outlier_flag=True,
                    algorithm_version="test-v1",
                ),
            ]
        )
        db.session.commit()

    response = client.get("/export/research.zip?niche=finance")

    assert response.status_code == 200
    assert response.mimetype == "application/zip"
    with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
        assert set(archive.namelist()) == {
            "channels.csv",
            "videos.csv",
            "manual_labels.csv",
            "snapshots.csv",
            "derived_metrics.csv",
            "content_theses.csv",
            "thesis_evidence.csv",
            "thesis_topics.csv",
            "thesis_scores.csv",
            "red_team_reviews.csv",
            "thesis_monetization_maps.csv",
            "sponsor_evidence.csv",
            "affiliate_product_evidence.csv",
            "assets.csv",
            "video_assets.csv",
            "video_rights_checklists.csv",
            "video_disclosures.csv",
            "owned_analytics_credentials.csv",
            "owned_video_analytics.csv",
            "retention_diagnostics.csv",
            "experiments.csv",
            "experiment_checkpoints.csv",
            "collection_runs.csv",
            "data_dictionary.md",
        }
        videos_csv = archive.read("videos.csv").decode()
        assert "youtube_video_id" in videos_csv
        assert "finance_video" in videos_csv
        assert "gaming_video" not in videos_csv

        labels_csv = archive.read("manual_labels.csv").decode()
        assert "entity_type" in labels_csv
        assert "finance" in labels_csv

        dictionary = archive.read("data_dictionary.md").decode()
        assert "Research Export Data Dictionary" in dictionary
        assert "outlier_flag" in dictionary
        assert "owned_video_analytics.csv" in dictionary


def test_research_jsonl_export_filters_outliers(client):
    with client.application.app_context():
        channel = Channel(channel_username="@jsonl_channel", subscribers=100)
        db.session.add(channel)
        db.session.flush()

        outlier = Video(
            youtube_video_id="jsonl_outlier",
            title="Outlier",
            views=1000,
            channel_id=channel.id,
        )
        normal = Video(
            youtube_video_id="jsonl_normal",
            title="Normal",
            views=100,
            channel_id=channel.id,
        )
        db.session.add_all([outlier, normal])
        db.session.flush()

        db.session.add_all(
            [
                VideoDerivedMetric(
                    video_id=outlier.id,
                    snapshot_at=datetime(2026, 1, 1, 0, 0, 0),
                    relative_performance=5,
                    outlier_flag=True,
                    algorithm_version="test-v1",
                ),
                VideoDerivedMetric(
                    video_id=normal.id,
                    snapshot_at=datetime(2026, 1, 1, 0, 0, 0),
                    relative_performance=1,
                    outlier_flag=False,
                    algorithm_version="test-v1",
                ),
            ]
        )
        db.session.commit()

    response = client.get("/export/research.jsonl?outlier_flag=true")

    assert response.status_code == 200
    assert response.mimetype == "application/x-ndjson"
    rows = [
        json.loads(line)
        for line in response.get_data(as_text=True).splitlines()
        if line.strip()
    ]
    derived_rows = [row for row in rows if row["dataset"] == "derived_metrics"]
    assert len(derived_rows) == 1
    assert derived_rows[0]["youtube_video_id"] == "jsonl_outlier"


def test_research_operations_pages_render_task_navigation(client):
    with client.application.app_context():
        channel = Channel(channel_username="@dashboard_channel", subscribers=1000)
        db.session.add(channel)
        db.session.flush()
        video = Video(
            youtube_video_id="dashboard_video",
            title="Dashboard Outlier",
            views=1000,
            channel_id=channel.id,
        )
        thesis = ContentThesis(
            thesis_id="DASH001",
            title="Dashboard thesis",
            status="research",
        )
        db.session.add_all([video, thesis])
        db.session.flush()
        db.session.add_all(
            [
                VideoLabel(video_id=video.id, niche="education"),
                VideoDerivedMetric(
                    video_id=video.id,
                    snapshot_at=datetime(2026, 1, 1, 0, 0, 0),
                    relative_performance=5,
                    outlier_flag=True,
                    algorithm_version="test-v1",
                ),
                CollectionRun(
                    run_type="channel_uploads",
                    status="completed",
                    input_type="channel_id",
                    input_value="UC_DASH",
                    quota_estimate=4,
                    items_saved=1,
                ),
            ]
        )
        db.session.commit()

    dashboard_response = client.get("/dashboard")
    assert dashboard_response.status_code == 200
    dashboard_body = dashboard_response.get_data(as_text=True)
    assert "Research Operations" in dashboard_body
    assert "Dashboard Outlier" in dashboard_body
    assert "Dashboard thesis" in dashboard_body

    for path, marker in (
        ("/collect", "Queue channel"),
        ("/exports", "Filtered Research ZIP"),
        ("/settings", "YouTube API key"),
        ("/operations", "Recent Failures"),
        ("/data?view=channels", "Enterprise Data Viewer"),
        ("/data?view=videos", "Enterprise Data Viewer"),
    ):
        response = client.get(path)
        assert response.status_code == 200
        assert marker in response.get_data(as_text=True)

    health_response = client.get("/healthz")
    assert health_response.status_code in {200, 503}
    health_payload = health_response.get_json()
    assert "database" in health_payload
    assert "redis" in health_payload
    assert health_payload["database"]["ok"] is True


def test_optional_admin_auth_protects_private_pages(monkeypatch):
    monkeypatch.setenv("DATABASE_URL", "sqlite:///:memory:")
    monkeypatch.setenv("ADMIN_PASSWORD", "correct-password")

    app = create_app()
    app.config["TESTING"] = True

    with app.app_context():
        db.drop_all()
        db.create_all()
        with app.test_client() as test_client:
            protected_response = test_client.get("/dashboard")
            assert protected_response.status_code == 302
            assert "/login" in protected_response.headers["Location"]

            health_response = test_client.get("/healthz")
            assert health_response.status_code in {200, 503}

            bad_login = test_client.post(
                "/login",
                data={"password": "wrong-password", "next": "/dashboard"},
                follow_redirects=True,
            )
            assert bad_login.status_code == 200
            assert "Invalid admin password." in bad_login.get_data(as_text=True)

            good_login = test_client.post(
                "/login",
                data={"password": "correct-password", "next": "/dashboard"},
                follow_redirects=True,
            )
            assert good_login.status_code == 200
            assert "Research Operations" in good_login.get_data(as_text=True)
        db.session.remove()
        db.drop_all()


def test_labeling_queue_displays_video_context(client):
    with client.application.app_context():
        channel = Channel(channel_username="@label_channel", subscribers=1500)
        db.session.add(channel)
        db.session.flush()
        video = Video(
            youtube_video_id="label_video_1",
            title="Label candidate",
            description_excerpt="Description for review",
            thumbnail_url="https://img.youtube.com/vi/label_video_1/hqdefault.jpg",
            views=5000,
            video_length="0:08:00",
            posted="2026-01-01",
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.commit()

    response = client.get("/labeling?mode=unlabeled")

    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "Label candidate" in body
    assert "@label_channel" in body
    assert "Description for review" in body
    assert "Open on YouTube" in body
    assert "faceless" in body


def test_save_video_label_creates_audit_and_rejects_invalid_vocab(client):
    with client.application.app_context():
        channel = Channel(channel_username="@audit_channel", subscribers=100)
        db.session.add(channel)
        db.session.flush()
        video = Video(
            youtube_video_id="audit_video",
            title="Audit candidate",
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.commit()
        video_id = video.id

    invalid_response = client.post(
        f"/labeling/{video_id}",
        data={
            "niche": "freeform typo",
            "format": "explainer",
            "faceless_status": "faceless",
            "ai_use_visible": "none_visible",
            "visual_style": "animation",
            "packaging_pattern": "how_to",
            "topic_type": "evergreen",
            "production_complexity": "low",
            "policy_risk": "low",
            "review_status": "reviewed",
            "reviewer": "reviewer-a",
            "label_confidence": "0.8",
        },
        follow_redirects=True,
    )
    assert invalid_response.status_code == 200

    with client.application.app_context():
        assert VideoLabel.query.count() == 0

    response = client.post(
        f"/labeling/{video_id}",
        data={
            "niche": "education",
            "format": "explainer",
            "faceless_status": "faceless",
            "ai_use_visible": "none_visible",
            "visual_style": "animation",
            "packaging_pattern": "how_to",
            "title_pattern": "specific_question",
            "thumbnail_pattern": "single_object_high_contrast",
            "curiosity_type": "mystery",
            "topic_type": "evergreen",
            "production_complexity": "low",
            "policy_risk": "low",
            "review_status": "reviewed",
            "reviewer": "reviewer-a",
            "label_confidence": "0.8",
            "viewer_promise": "Understand the surprising system.",
            "clarity_score": "5",
            "specificity_score": "4",
            "honesty_score": "5",
            "visual_readability_score": "4",
            "differentiation_score": "3",
            "monetization_signals": "Sponsor fit",
            "notes": "Strong packaging",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    with client.application.app_context():
        label = VideoLabel.query.one()
        assert label.niche == "education"
        assert label.label_confidence == 0.8
        assert label.title_pattern == "specific_question"
        assert label.thumbnail_pattern == "single_object_high_contrast"
        assert label.viewer_promise == "Understand the surprising system."
        assert label.clarity_score == 5
        audit = VideoLabelAudit.query.one()
        assert audit.action == "reviewed"
        assert audit.reviewer == "reviewer-a"
        assert audit.previous_values == {}
        assert audit.new_values["niche"] == "education"
        assert audit.new_values["title_pattern"] == "specific_question"


def test_bulk_label_applies_controlled_value_to_selected_videos(client):
    with client.application.app_context():
        channel = Channel(channel_username="@bulk_channel", subscribers=100)
        db.session.add(channel)
        db.session.flush()
        videos = [
            Video(
                youtube_video_id=f"bulk_{index}",
                title=f"Bulk {index}",
                channel_id=channel.id,
            )
            for index in range(3)
        ]
        db.session.add_all(videos)
        db.session.commit()
        selected_ids = [videos[0].id, videos[1].id]

    response = client.post(
        "/labeling/bulk",
        data={
            "video_ids": [str(video_id) for video_id in selected_ids],
            "field": "policy_risk",
            "value": "medium",
            "reviewer": "bulk-reviewer",
        },
        follow_redirects=True,
    )

    assert response.status_code == 200
    with client.application.app_context():
        labels = VideoLabel.query.order_by(VideoLabel.video_id.asc()).all()
        assert len(labels) == 2
        assert {label.policy_risk for label in labels} == {"medium"}
        assert VideoLabelAudit.query.count() == 2


def test_market_analysis_compute_route_creates_derived_metrics(client):
    with client.application.app_context():
        channel = Channel(channel_username="@analysis_channel", subscribers=1000)
        db.session.add(channel)
        db.session.flush()
        baseline = Video(
            youtube_video_id="analysis_baseline",
            title="Baseline",
            views=1000,
            likes=20,
            comments=5,
            duration_seconds=600,
            published_at=datetime(2026, 1, 1, 0, 0, 0),
            channel_id=channel.id,
        )
        outlier = Video(
            youtube_video_id="analysis_outlier",
            title="Analysis Outlier",
            views=3000,
            likes=300,
            comments=30,
            duration_seconds=600,
            published_at=datetime(2026, 1, 2, 0, 0, 0),
            channel_id=channel.id,
        )
        db.session.add_all([baseline, outlier])
        db.session.flush()
        db.session.add_all(
            [
                VideoSnapshot(
                    video_id=baseline.id,
                    snapshot_at=datetime(2026, 1, 10, 0, 0, 0),
                    view_count=1000,
                    like_count=20,
                    comment_count=5,
                    subscriber_count_at_snapshot=1000,
                ),
                VideoSnapshot(
                    video_id=outlier.id,
                    snapshot_at=datetime(2026, 1, 10, 0, 0, 0),
                    view_count=3000,
                    like_count=300,
                    comment_count=30,
                    subscriber_count_at_snapshot=1000,
                ),
                VideoLabel(
                    video_id=outlier.id,
                    niche="education",
                    format="explainer",
                    topic_type="evergreen",
                    packaging_pattern="how_to",
                    review_status="reviewed",
                ),
            ]
        )
        db.session.commit()

    response = client.post("/analysis/compute", follow_redirects=True)

    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "Market Opportunity View" in body
    assert "Analysis Outlier" in body
    assert "Repeated Outlier Topics" in body
    assert "Under-served Candidate Theses" in body
    assert "education" in body
    with client.application.app_context():
        assert VideoDerivedMetric.query.count() == 2
        assert ChannelDerivedSummary.query.count() == 1


def test_packaging_lab_displays_patterns_changes_and_saves_experiment(client):
    with client.application.app_context():
        channel = Channel(channel_username="@packaging_channel", subscribers=1000)
        db.session.add(channel)
        db.session.flush()
        video = Video(
            youtube_video_id="packaging_video",
            title="Why This System Works",
            views=5000,
            thumbnail_url="https://img.youtube.com/vi/packaging_video/hqdefault.jpg",
            thumbnail_quality="high",
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.flush()
        db.session.add_all(
            [
                VideoLabel(
                    video_id=video.id,
                    niche="education",
                    format="explainer",
                    title_pattern="specific_question",
                    thumbnail_pattern="single_object_high_contrast",
                    viewer_promise="Learn the hidden system.",
                    curiosity_type="hidden_system",
                    clarity_score=5,
                    specificity_score=4,
                    honesty_score=5,
                    visual_readability_score=4,
                    differentiation_score=3,
                    review_status="reviewed",
                ),
                VideoDerivedMetric(
                    video_id=video.id,
                    snapshot_at=datetime(2026, 1, 5, 0, 0, 0),
                    channel_recent_median_views=1000,
                    relative_performance=5,
                    outlier_flag=True,
                    algorithm_version="test-v1",
                ),
                VideoSnapshot(
                    video_id=video.id,
                    snapshot_at=datetime(2026, 1, 4, 0, 0, 0),
                    view_count=1000,
                ),
                VideoSnapshot(
                    video_id=video.id,
                    snapshot_at=datetime(2026, 1, 6, 0, 0, 0),
                    view_count=5000,
                ),
                VideoMetadataChange(
                    video_id=video.id,
                    field_name="title",
                    old_value="Old Title",
                    new_value="Why This System Works",
                    changed_at=datetime(2026, 1, 5, 0, 0, 0),
                ),
            ]
        )
        db.session.commit()

    response = client.get("/packaging?niche=education")
    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "Packaging Lab" in body
    assert "Why This System Works" in body
    assert "single_object_high_contrast" in body
    assert "Metadata Change Analysis" in body
    assert "4000" in body

    create_response = client.post(
        "/packaging/experiments",
        data={
            "working_title": "Pilot video",
            "niche": "education",
            "format": "explainer",
            "title_candidates": "Title A\nTitle B",
            "thumbnail_concepts": "Object on white\nDiagram cue",
            "experiment_log_url": "https://example.com/log",
            "final_title": "Title A",
            "final_thumbnail_concept": "Object on white",
            "final_choice_reason": "Highest clarity.",
            "status": "selected",
        },
        follow_redirects=True,
    )
    assert create_response.status_code == 200
    assert "Pilot video" in create_response.get_data(as_text=True)
    with client.application.app_context():
        experiment = PackagingExperiment.query.one()
        assert experiment.title_candidates == ["Title A", "Title B"]
        assert experiment.thumbnail_concepts == ["Object on white", "Diagram cue"]
        assert experiment.final_thumbnail_concept == "Object on white"
        assert experiment.status == "selected"


def test_thesis_workspace_creates_scores_evidence_topics_and_red_team_review(client):
    with client.application.app_context():
        channel = Channel(channel_username="@thesis_channel", subscribers=1000)
        db.session.add(channel)
        db.session.flush()
        video = Video(
            youtube_video_id="thesis_video",
            title="Thesis evidence video",
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.commit()
        channel_id = channel.id
        video_id = video.id

    create_response = client.post(
        "/theses",
        data={
            "thesis_id": "T001",
            "title": "Hidden systems explainers",
            "target_viewer": "Curious adult learners",
            "viewer_promise": "Understand overlooked systems.",
            "format": "explainer",
            "topic_universe": "Forgotten infrastructure and companies.",
            "production_edge": "Primary-source research.",
            "packaging_edge": "Specific question titles.",
            "monetization_path": "Ads and software sponsors.",
            "policy_risk_argument": "Original scripts and commentary.",
            "status": "research",
            "notes": "Promising candidate.",
        },
        follow_redirects=True,
    )
    assert create_response.status_code == 200
    assert "Hidden systems explainers" in create_response.get_data(as_text=True)

    with client.application.app_context():
        thesis = ContentThesis.query.one()
        thesis_id = thesis.id

    blocked_launch = client.post(
        f"/theses/{thesis_id}/status",
        data={"status": "launch"},
        follow_redirects=True,
    )
    assert blocked_launch.status_code == 200
    assert "cannot move to launch without a monetization map" in (
        blocked_launch.get_data(as_text=True)
    )

    client.post(
        f"/theses/{thesis_id}/evidence",
        data={
            "evidence_type": "outlier_video",
            "channel_id": str(channel_id),
            "video_id": str(video_id),
            "source_url": "https://youtube.com/watch?v=thesis_video",
            "note": "Breakout competitor proof.",
            "confidence": "0.8",
        },
        follow_redirects=True,
    )
    client.post(
        f"/theses/{thesis_id}/topics",
        data={
            "topic": "Why old payment rails still exist",
            "title_angle": "Why does this still exist?",
            "demand_evidence": "Repeated outliers.",
            "source_availability": "high",
            "production_complexity": "medium",
            "packaging_potential": "high",
            "status": "shortlisted",
        },
        follow_redirects=True,
    )
    client.post(
        f"/theses/{thesis_id}/scores",
        data={
            "factor": "audience_demand",
            "score": "5",
            "evidence": "Multiple outliers.",
            "confidence": "0.9",
        },
        follow_redirects=True,
    )
    client.post(
        f"/theses/{thesis_id}/monetization",
        data={
            "revenue_paths": ["watch_page_ads", "sponsors", "affiliates"],
            "primary_revenue_path": "watch_page_ads",
            "secondary_revenue_path": "sponsors",
            "conservative_ad_rpm": "2.5",
            "base_ad_rpm": "5.0",
            "upside_ad_rpm": "9.0",
            "sponsor_rpm_equivalent": "4.5",
            "affiliate_rpm_equivalent": "1.5",
            "membership_rpm_equivalent": "0.5",
            "product_rpm_equivalent": "0.75",
            "break_even_view_count": "25000",
            "meaningful_income_view_count": "250000",
            "assumptions": "Ads, sponsors, and affiliates modeled separately.",
            "main_monetization_risk": "Sponsor scale may arrive late.",
        },
        follow_redirects=True,
    )
    client.post(
        f"/theses/{thesis_id}/sponsor-evidence",
        data={
            "sponsor_category": "SaaS",
            "observed_sponsor": "ExampleSponsor",
            "competitor_channel_id": str(channel_id),
            "video_url": "https://youtube.com/watch?v=thesis_video",
            "date_observed": "2026-01-07",
            "niche_fit": "high",
            "brand_safety_notes": "Advertiser-safe education framing.",
        },
        follow_redirects=True,
    )
    client.post(
        f"/theses/{thesis_id}/affiliate-evidence",
        data={
            "product_category": "research tools",
            "program_source": "Example affiliate program",
            "estimated_fit": "medium",
            "audience_intent": "learn and buy better tools",
            "compliance_disclosure_concerns": "Needs clear disclosure.",
        },
        follow_redirects=True,
    )
    red_team_response = client.post(
        f"/theses/{thesis_id}/red-team",
        data={
            "reviewer": "reviewer-a",
            "decision_under_review": "pilot",
            "decision": "proceed_to_pilot",
            "better_channels_answer": "Existing channels are broad.",
            "weak_monetization_answer": "Sponsor categories are visible.",
            "failure_premortem": "The channel failed from weak packaging.",
            "early_warning_signs": "Low CTR.",
            "preventive_actions": "Pre-test thumbnails.",
            "kill_criteria": "Three pilots under baseline.",
            "competitor_challenges": "Competitor A | better archives",
            "decision_rationale": "Evidence supports a small pilot.",
        },
        follow_redirects=True,
    )
    assert red_team_response.status_code == 200
    assert "proceed to pilot" in red_team_response.get_data(as_text=True)

    status_response = client.post(
        f"/theses/{thesis_id}/status",
        data={"status": "launch"},
        follow_redirects=True,
    )
    assert status_response.status_code == 200

    with client.application.app_context():
        thesis = ContentThesis.query.one()
        assert thesis.status == "launch"
        assert ThesisEvidence.query.one().confidence == 0.8
        assert ThesisTopic.query.one().status == "shortlisted"
        assert ThesisScore.query.one().weighted_score == 25
        monetization_map = ThesisMonetizationMap.query.one()
        assert monetization_map.revenue_paths == [
            "watch_page_ads",
            "sponsors",
            "affiliates",
        ]
        assert monetization_map.base_ad_rpm == 5.0
        assert SponsorEvidence.query.one().observed_sponsor == "ExampleSponsor"
        assert AffiliateProductEvidence.query.one().product_category == "research tools"
        review = RedTeamReview.query.one()
        assert review.core_objections["better_channels"]["answer"] == (
            "Existing channels are broad."
        )

    export_response = client.get("/export/research.jsonl")
    assert export_response.status_code == 200
    rows = [
        json.loads(line)
        for line in export_response.get_data(as_text=True).splitlines()
        if line.strip()
    ]
    assert any(row["dataset"] == "content_theses" for row in rows)
    assert any(row["dataset"] == "red_team_reviews" for row in rows)
    assert any(row["dataset"] == "thesis_monetization_maps" for row in rows)
    assert any(row["dataset"] == "sponsor_evidence" for row in rows)
    assert any(row["dataset"] == "affiliate_product_evidence" for row in rows)


def test_rights_workspace_blocks_unclear_assets_and_exports_records(client):
    with client.application.app_context():
        channel = Channel(channel_username="@rights_channel", subscribers=100)
        db.session.add(channel)
        db.session.flush()
        video = Video(
            youtube_video_id="rights_video",
            title="Rights candidate",
            channel_id=channel.id,
        )
        safe_video = Video(
            youtube_video_id="rights_video_safe",
            title="Rights safe candidate",
            channel_id=channel.id,
        )
        db.session.add_all([video, safe_video])
        db.session.commit()
        video_id = video.id
        safe_video_id = safe_video.id

    high_risk_response = client.post(
        "/rights/assets",
        data={
            "video_id": str(video_id),
            "asset_id": "A001",
            "asset_type": "music",
            "source_url_path": "https://example.com/free-music",
            "creator_licensor": "Unknown",
            "license_terms": "Unclear royalty-free page.",
            "monetized_youtube_allowed": "unclear",
            "high_risk_flag": "on",
            "high_risk_reason": "Unknown royalty-free music.",
        },
        follow_redirects=True,
    )
    assert high_risk_response.status_code == 200
    assert "A001" in high_risk_response.get_data(as_text=True)

    with client.application.app_context():
        risky_asset = Asset.query.filter_by(asset_id="A001").one()
        assert risky_asset.high_risk_flag is True
        risky_asset_id = risky_asset.id

    client.post(
        "/rights/video-assets",
        data={
            "video_id": str(video_id),
            "asset_id": str(risky_asset_id),
            "intended_use": "Background music.",
            "rights_decision": "use",
        },
        follow_redirects=True,
    )
    blocked_ready = client.post(
        f"/rights/{video_id}/checklists",
        data={
            "every_asset_has_row": "on",
            "unclear_assets_blocked": "on",
            "attribution_captured": "on",
            "no_terms_prohibit_monetization": "on",
            "ready_for_upload": "on",
            "synthetic_altered_status": "none",
            "reviewer": "rights-reviewer",
        },
        follow_redirects=True,
    )
    assert "Blocked or unproven assets prevent upload" in blocked_ready.get_data(
        as_text=True
    )

    client.post(
        "/rights/assets",
        data={
            "video_id": str(safe_video_id),
            "asset_id": "A002",
            "asset_type": "original_graphic",
            "source_url_path": "/assets/diagram.png",
            "creator_licensor": "internal",
            "license_terms": "Owned original graphic.",
            "monetized_youtube_allowed": "yes",
            "proof_saved": "on",
        },
        follow_redirects=True,
    )
    with client.application.app_context():
        safe_asset_id = Asset.query.filter_by(asset_id="A002").one().id

    client.post(
        "/rights/video-assets",
        data={
            "video_id": str(safe_video_id),
            "asset_id": str(safe_asset_id),
            "intended_use": "Main diagram.",
            "rights_decision": "use",
        },
        follow_redirects=True,
    )
    ready_response = client.post(
        f"/rights/{safe_video_id}/checklists",
        data={
            "every_asset_has_row": "on",
            "unclear_assets_blocked": "on",
            "attribution_captured": "on",
            "no_terms_prohibit_monetization": "on",
            "ready_for_upload": "on",
            "synthetic_altered_status": "none",
            "reviewer": "rights-reviewer",
        },
        follow_redirects=True,
    )
    assert ready_response.status_code == 200

    disclosure_response = client.post(
        f"/rights/{safe_video_id}/disclosures",
        data={
            "sponsor_disclosure": "No sponsor.",
            "affiliate_disclosure": "Affiliate links disclosed.",
            "altered_synthetic_disclosure": "No synthetic media.",
            "music_license_attribution": "Original music only.",
            "disclosure_notes": "Ready for audit.",
        },
        follow_redirects=True,
    )
    assert disclosure_response.status_code == 200

    with client.application.app_context():
        assert VideoAsset.query.count() == 2
        assert VideoRightsChecklist.query.count() == 1
        assert VideoRightsChecklist.query.one().ready_for_upload is True
        assert VideoDisclosure.query.one().affiliate_disclosure == (
            "Affiliate links disclosed."
        )

    export_response = client.get("/export/research.jsonl")
    rows = [
        json.loads(line)
        for line in export_response.get_data(as_text=True).splitlines()
        if line.strip()
    ]
    assert any(row["dataset"] == "assets" for row in rows)
    assert any(row["dataset"] == "video_rights_checklists" for row in rows)


def test_owned_analytics_workspace_records_private_metrics_and_experiments(client):
    with client.application.app_context():
        channel = Channel(
            channel_username="@owned_channel",
            youtube_channel_id="UC_OWNED",
            subscribers=100,
        )
        db.session.add(channel)
        db.session.flush()
        video = Video(
            youtube_video_id="owned_video",
            title="Owned pilot",
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.commit()
        channel_id = channel.id
        video_id = video.id

    page = client.get(f"/owned?video_id={video_id}")
    assert page.status_code == 200
    assert "Owned Analytics" in page.get_data(as_text=True)
    assert "Competitor rows stay limited to public YouTube data." in page.get_data(
        as_text=True
    )

    credential_response = client.post(
        "/owned/credentials",
        data={
            "channel_id": str(channel_id),
            "google_account_email": "owner@example.com",
            "token_secret_ref": "secret-manager://youtube/owned-channel",
            "notes": "Authorized test channel.",
        },
        follow_redirects=True,
    )
    assert credential_response.status_code == 200

    with client.application.app_context():
        credential = OwnedAnalyticsCredential.query.one()
        assert credential.status == "configured"
        assert credential.token_secret_ref == "secret-manager://youtube/owned-channel"
        credential_id = credential.id

    revoke_response = client.post(
        f"/owned/credentials/{credential_id}/revoke",
        follow_redirects=True,
    )
    assert revoke_response.status_code == 200

    client.post(
        "/owned/analytics",
        data={
            "video_id": str(video_id),
            "date": "2026-05-11",
            "views": "1000",
            "impressions": "25000",
            "impression_ctr": "4.0",
            "average_view_duration_seconds": "360",
            "average_view_percentage": "52",
            "watch_time_minutes": "6000",
            "subscribers_gained": "25",
            "estimated_revenue": "12.50",
            "traffic_source_type": "browse",
            "source": "manual",
        },
        follow_redirects=True,
    )
    client.post(
        "/owned/retention",
        data={
            "video_id": str(video_id),
            "report_date": "2026-05-11",
            "ctr": "4.0",
            "average_view_duration_seconds": "360",
            "average_view_percentage": "52",
            "impressions": "25000",
            "dominant_traffic_source": "browse",
            "retention_pattern": "high_ctr_low_retention",
            "likely_cause": "Opening promise is too broad.",
            "evidence": "Drop before first proof point.",
            "next_change": "Rewrite first 30 seconds.",
        },
        follow_redirects=True,
    )
    experiment_response = client.post(
        "/owned/experiments",
        data={
            "video_id": str(video_id),
            "hypothesis": "Sharper title improves qualified CTR.",
            "variable_tested": "title",
            "title": "Why This System Breaks",
            "thumbnail_variant": "variant-a",
            "publish_date": "2026-05-11",
            "success_metric": "7d APV over 50",
            "production_hours": "8",
            "production_cost": "25",
            "decision": "pending",
            "notes": "Pilot test.",
        },
        follow_redirects=True,
    )
    assert experiment_response.status_code == 200

    with client.application.app_context():
        experiment = Experiment.query.one()
        experiment_id = experiment.id

    client.post(
        f"/owned/experiments/{experiment_id}/checkpoints",
        data={
            "checkpoint": "24h",
            "views": "1000",
            "impressions": "25000",
            "impression_ctr": "4.0",
            "average_view_duration_seconds": "360",
            "average_view_percentage": "52",
            "watch_time_minutes": "6000",
            "subscribers_gained": "25",
            "main_traffic_source": "browse",
            "notes": "Healthy first checkpoint.",
        },
        follow_redirects=True,
    )

    with client.application.app_context():
        assert OwnedAnalyticsCredential.query.one().status == "revoked"
        assert OwnedVideoAnalytics.query.one().estimated_revenue == 12.50
        assert RetentionDiagnostic.query.one().retention_pattern == (
            "high_ctr_low_retention"
        )
        assert ExperimentCheckpoint.query.one().checkpoint == "24h"

    export_response = client.get("/export/research.jsonl")
    rows = [
        json.loads(line)
        for line in export_response.get_data(as_text=True).splitlines()
        if line.strip()
    ]
    assert any(row["dataset"] == "owned_video_analytics" for row in rows)
    assert any(row["dataset"] == "retention_diagnostics" for row in rows)
    assert any(row["dataset"] == "experiments" for row in rows)
    assert any(row["dataset"] == "experiment_checkpoints" for row in rows)


def test_video_detail_route_success(client):
    with client.application.app_context():
        channel = Channel(channel_username="@video_detail_channel", subscribers=1200)
        db.session.add(channel)
        db.session.flush()

        video = Video(
            youtube_video_id="video_detail_123",
            title="Video detail test",
            views=100,
            likes=10,
            comments=1,
            posted="2025-01-01",
            video_length="5:00",
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.commit()
        video_id = video.id

    response = client.get(f"/video/{video_id}")

    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "Video detail test" in body


def test_video_history_api_returns_ordered_time_series(client):
    with client.application.app_context():
        channel = Channel(channel_username="@history_channel", subscribers=2500)
        db.session.add(channel)
        db.session.flush()

        video = Video(
            youtube_video_id="video_history_123",
            title="History API test",
            views=100,
            likes=12,
            comments=2,
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.flush()

        db.session.add_all(
            [
                VideoHistory(
                    video_id=video.id,
                    views=50,
                    likes=5,
                    comments=1,
                    timestamp=datetime(2026, 1, 1, 0, 0, 0),
                ),
                VideoHistory(
                    video_id=video.id,
                    views=100,
                    likes=12,
                    comments=2,
                    timestamp=datetime(2026, 1, 2, 0, 0, 0),
                ),
            ]
        )
        db.session.commit()
        video_id = video.id

    response = client.get(f"/api/video/{video_id}/history")
    assert response.status_code == 200

    payload = response.get_json()
    assert payload["timestamps"] == ["2026-01-01T00:00:00Z", "2026-01-02T00:00:00Z"]
    assert payload["views"] == [50, 100]
    assert payload["likes"] == [5, 12]
    assert payload["comments"] == [1, 2]


def test_refresh_video_detail_updates_stats_and_appends_history(client, monkeypatch):
    monkeypatch.setattr(routes, "YOUTUBE_API_KEY", "test-api-key")

    with client.application.app_context():
        channel = Channel(channel_username="@refresh_channel", subscribers=1200)
        db.session.add(channel)
        db.session.flush()

        video = Video(
            youtube_video_id="refresh_video_1",
            title="Before refresh",
            views=100,
            likes=10,
            comments=1,
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.commit()
        video_id = video.id

    monkeypatch.setattr(
        routes,
        "get_video_data",
        lambda _video_id: {
            "youtube_video_id": "refresh_video_1",
            "channel_username": "@refresh_channel",
            "subscribers": "1250",
            "title": "After refresh",
            "description": "Updated description",
            "views": "9999",
            "likes": "350",
            "comments": "40",
            "posted": "2026-01-01",
            "video_length": "0:12:00",
            "transcript": "Updated transcript",
        },
    )

    response = client.post(f"/video/{video_id}/refresh", follow_redirects=True)
    assert response.status_code == 200

    with client.application.app_context():
        refreshed_video = Video.query.get(video_id)
        assert refreshed_video.title == "After refresh"
        assert refreshed_video.views == 9999
        assert refreshed_video.likes == 350
        assert refreshed_video.comments == 40

        history_rows = (
            VideoHistory.query.filter_by(video_id=video_id)
            .order_by(VideoHistory.timestamp.asc())
            .all()
        )
        assert len(history_rows) == 1
        assert history_rows[0].views == 9999
        assert history_rows[0].likes == 350
        assert history_rows[0].comments == 40


def test_video_like_rate_bdd_scenario(client):
    with client.application.app_context():
        channel = Channel(channel_username="@engagement_channel", subscribers=8000)
        db.session.add(channel)
        db.session.flush()

        video = Video(
            youtube_video_id="engagement_1",
            title="Engagement baseline",
            views=10000,
            likes=500,
            comments=200,
            channel_id=channel.id,
        )
        db.session.add(video)
        db.session.commit()
        video_id = video.id

        assert video.like_rate == 5.0
        assert video.comment_rate == 2.0
        assert video.engagement_rate == 7.0

    response = client.get(f"/video/{video_id}")
    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "5.00%" in body
    assert "2.00%" in body
    assert "7.00%" in body


def test_video_engagement_handles_zero_views_and_none(client):
    with client.application.app_context():
        channel = Channel(channel_username="@zero_metrics_channel", subscribers=900)
        db.session.add(channel)
        db.session.flush()

        zero_video = Video(
            youtube_video_id="engagement_zero",
            title="Zero Engagement",
            views=0,
            likes=0,
            comments=0,
            channel_id=channel.id,
        )
        none_video = Video(
            youtube_video_id="engagement_none",
            title="None Engagement",
            views=None,
            likes=None,
            comments=None,
            channel_id=channel.id,
        )
        db.session.add_all([zero_video, none_video])
        db.session.commit()
        zero_video_id = zero_video.id

        assert zero_video.like_rate == 0.0
        assert zero_video.comment_rate == 0.0
        assert zero_video.engagement_rate == 0.0
        assert none_video.like_rate == 0.0
        assert none_video.comment_rate == 0.0
        assert none_video.engagement_rate == 0.0

    response = client.get(f"/video/{zero_video_id}")
    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "0.00%" in body


def test_channel_detail_route_success(client):
    with client.application.app_context():
        channel = Channel(channel_username="@channel_detail_channel", subscribers=5000)
        db.session.add(channel)
        db.session.flush()

        video = Video(
            youtube_video_id="channel_detail_video_1",
            title="Linked channel video",
            views=250,
            likes=25,
            comments=3,
            channel_id=channel.id,
        )
        history = ChannelHistory(
            channel_id=channel.id,
            previous_subscribers=4900,
        )
        db.session.add_all([video, history])
        db.session.commit()
        channel_id = channel.id

    response = client.get(f"/channel/{channel_id}")

    assert response.status_code == 200
    body = response.get_data(as_text=True)
    assert "@channel_detail_channel" in body
    assert "Linked channel video" in body
    assert "4,900" in body


def test_toggle_channel_tracking_api(client):
    with client.application.app_context():
        channel = Channel(channel_username="@toggle_tracking_channel", subscribers=750)
        db.session.add(channel)
        db.session.commit()
        channel_id = channel.id

    response = client.post(f"/api/channel/{channel_id}/toggle-tracking")
    assert response.status_code == 200
    assert response.get_json() == {"is_tracked": True}

    with client.application.app_context():
        channel = Channel.query.get(channel_id)
        assert channel.is_tracked is True

    response = client.post(f"/api/channel/{channel_id}/toggle-tracking")
    assert response.status_code == 200
    assert response.get_json() == {"is_tracked": False}

    with client.application.app_context():
        channel = Channel.query.get(channel_id)
        assert channel.is_tracked is False
